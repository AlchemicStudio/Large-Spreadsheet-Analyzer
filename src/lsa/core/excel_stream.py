"""Row streams for XLS/XLSX/ODS workbooks.

XLSX is read with openpyxl in read-only mode, which truly streams rows in
constant memory.  XLS and ODS are read with python-calamine: fast, but the
sheet's used range is materialized by the underlying Rust library, so very
large XLS/ODS files cost memory proportional to their size — callers should
warn users about this (see :data:`FORMAT_MEMORY_WARNING`).
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook
from python_calamine import CalamineWorkbook

from .rows import FileFormat, Row, cell_to_text, detect_format, suppress_trailing_empty

FORMAT_MEMORY_WARNING = (
    "XLS and ODS files cannot be read row-by-row; the whole sheet is loaded "
    "in memory. Very large files may be slow or exhaust RAM — consider "
    "converting to CSV or XLSX first."
)


class SheetNotFoundError(ValueError):
    """The requested sheet does not exist in the workbook."""


def _pick_sheet(requested: str | None, names: list[str]) -> str:
    if not names:
        raise SheetNotFoundError("the workbook contains no sheets")
    if requested is None:
        return names[0]
    if requested not in names:
        available = ", ".join(repr(n) for n in names)
        raise SheetNotFoundError(f"no sheet named {requested!r} (available: {available})")
    return requested


def list_sheets(path: str | Path) -> list[str]:
    """Return the sheet names of an XLS/XLSX/ODS workbook, in file order."""
    return CalamineWorkbook.from_path(str(path)).sheet_names


class _ExcelStreamBase:
    """Shared peek/numbering logic for workbook-based streams."""

    supports_offsets = False

    header: list[str] | None
    width: int
    total_rows: int | None
    _has_header: bool
    _pending: Row | None

    def _init_layout(self, first: list[str] | None, has_header: bool) -> None:
        self._has_header = has_header
        self.header = None
        self._pending = None
        if first is None:
            self.width = 0
        elif has_header:
            self.header = first
            self.width = len(first)
        else:
            self.width = len(first)
            self._pending = Row(1, first, None)

    def _set_total_rows(self, physical_rows: int | None) -> None:
        """Store the *data*-row estimate (header excluded) for progress."""
        if physical_rows is None:
            self.total_rows = None
        else:
            self.total_rows = max(0, physical_rows - (1 if self._has_header else 0))

    def _numbered(self, raw_rows: Iterator[list[str]]) -> Iterator[Row]:
        number = 2 if self._has_header else 1
        if self._pending is not None:
            pending, self._pending = self._pending, None
            yield pending
            number = pending.number + 1
        for cells in raw_rows:
            yield Row(number, cells, None)
            number += 1

    def bytes_read(self) -> None:
        """Workbook formats report progress in rows, not bytes."""
        return None

    def total_bytes(self) -> None:
        return None

    def __enter__(self) -> _ExcelStreamBase:
        return self

    def __exit__(self, *exc_info: object) -> None:
        self.close()  # type: ignore[attr-defined]


class CalamineRowStream(_ExcelStreamBase):
    """RowStream over XLS/ODS (any calamine-supported workbook, in fact)."""

    def __init__(self, path: str | Path, *, sheet: str | None = None, has_header: bool = True):
        self._wb = CalamineWorkbook.from_path(str(path))
        name = _pick_sheet(sheet, self._wb.sheet_names)
        ws = self._wb.get_sheet_by_name(name)
        self.sheet_name = name
        # iter_rows() yields every physical row from row 1 but trims leading
        # empty *columns* to the used range; pad them back so letter/index
        # references keep their absolute meaning.
        start = ws.start
        self._col_pad = [""] * (start[1] if start is not None else 0)
        if start is None:
            # Empty sheet: calamine 0.8's iter_rows() panics (a PanicException
            # that no `except Exception` can catch) when there is no used
            # range — never call it in that case.
            self._rows_iter: Iterator[list[object]] = iter(())
        else:
            self._rows_iter = iter(ws.iter_rows())
        first_raw = next(self._rows_iter, None)
        first = None if first_raw is None else self._to_text(first_raw)
        self._init_layout(first, has_header)
        self._set_total_rows((start[0] if start is not None else 0) + ws.height)

    def _to_text(self, raw: list[object]) -> list[str]:
        return self._col_pad + [cell_to_text(v) for v in raw]

    def rows(self) -> Iterator[Row]:
        raw = (self._to_text(r) for r in self._rows_iter)
        return suppress_trailing_empty(self._numbered(raw))

    def close(self) -> None:
        self._wb.close()


class OpenpyxlRowStream(_ExcelStreamBase):
    """RowStream over XLSX/XLSM via openpyxl read-only mode (true streaming)."""

    def __init__(self, path: str | Path, *, sheet: str | None = None, has_header: bool = True):
        self._wb = load_workbook(str(path), read_only=True, data_only=True)
        try:
            name = _pick_sheet(sheet, self._wb.sheetnames)
            ws = self._wb[name]
            self.sheet_name = name
            physical_estimate = ws.max_row  # progress estimate only
            # In read-only mode the stated <dimension> BOUNDS iteration: a
            # writer that lies (the classic `<dimension ref="A1"/>` from
            # streaming generators, or a stale range) would silently truncate
            # rows and columns.  Always recalculate from the actual cells and
            # keep the stated value only as a progress estimate.
            ws.reset_dimensions()
            self._rows_iter: Iterator[tuple[object, ...]] = ws.iter_rows(values_only=True)
            first_raw = next(self._rows_iter, None)
            first = None if first_raw is None else [cell_to_text(v) for v in first_raw]
            self._init_layout(first, has_header)
            self._set_total_rows(physical_estimate)
        except BaseException:
            self._wb.close()
            raise

    def rows(self) -> Iterator[Row]:
        raw = ([cell_to_text(v) for v in r] for r in self._rows_iter)
        return suppress_trailing_empty(self._numbered(raw))

    def close(self) -> None:
        self._wb.close()


def needs_memory_warning(path: str | Path, *, threshold_bytes: int = 50 * 1024 * 1024) -> bool:
    """True when the file is a large XLS/ODS that will be fully materialized."""
    fmt = detect_format(path)
    if fmt not in (FileFormat.XLS, FileFormat.ODS):
        return False
    return Path(path).stat().st_size >= threshold_bytes
